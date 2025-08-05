import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from typing import Dict, List, Any, Tuple, Optional
import io
import base64
from matplotlib.figure import Figure
import warnings
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
import os
warnings.filterwarnings('ignore')

try:
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
    QT_AVAILABLE = True
except ImportError:
    QT_AVAILABLE = False
    FigureCanvas = None


class BusinessDataQualityAnalyzer:
    """Data quality analysis with clear, actionable insights."""
    
    def __init__(self, df: pd.DataFrame):
        self.df = df.copy()
        self.report = {}
        self.business_issues = []
        self._generate_business_report()
    
    def _generate_business_report(self):
        """To generate business-friendly data quality report."""
        # Generate reports in order to avoid circular dependencies
        self.report = {}
        self.report['overview'] = self._get_business_overview()
        self.report['completeness'] = self._analyze_data_completeness()
        self.report['consistency'] = self._analyze_data_consistency()
        self.report['accuracy'] = self._analyze_data_accuracy()
        self.report['reliability'] = self._analyze_data_reliability()
        self.report['business_impact'] = self._assess_business_impact()
    
    def _get_business_overview(self) -> Dict[str, Any]:
        """Get business-friendly dataset overview."""
        total_cells = len(self.df) * len(self.df.columns)
        missing_cells = self.df.isnull().sum().sum()
        complete_cells = total_cells - missing_cells
        
        # Calculate data health score
        completeness_score = (complete_cells / total_cells) * 100 if total_cells > 0 else 0
        
        return {
            'total_records': len(self.df),
            'total_fields': len(self.df.columns),
            'data_health_score': round(completeness_score, 1),
            'complete_records': len(self.df) - self.df.isnull().any(axis=1).sum(),
            'incomplete_records': self.df.isnull().any(axis=1).sum(),
            'data_size_mb': round(self.df.memory_usage(deep=True).sum() / 1024 / 1024, 2),
            'health_status': self._get_health_status(completeness_score)
        }
    
    def _get_health_status(self, score: float) -> str:
        """Get business-friendly health status."""
        if score >= 95:
            return "Excellent - Ready for analysis"
        elif score >= 85:
            return "Good - Minor cleanup needed"
        elif score >= 70:
            return "Fair - Some attention required"
        elif score >= 50:
            return "Poor - Significant cleanup needed"
        else:
            return "Critical - Major data issues found"
    
    def _analyze_data_completeness(self) -> Dict[str, Any]:
        """Analyze data completeness in business terms."""
        missing_counts = self.df.isnull().sum()
        missing_percentages = (missing_counts / len(self.df)) * 100
        
        # Categorize fields by completeness
        excellent_fields = []  # < 5% missing
        good_fields = []       # 5-15% missing
        concerning_fields = []  # 15-50% missing
        critical_fields = []   # > 50% missing
        
        for col in self.df.columns:
            missing_pct = missing_percentages[col]
            if missing_pct < 5:
                excellent_fields.append(col)
            elif missing_pct < 15:
                good_fields.append(col)
            elif missing_pct < 50:
                concerning_fields.append(col)
            else:
                critical_fields.append(col)
        
        # Business recommendations
        recommendations = []
        if critical_fields:
            recommendations.append(f"⚠️ {len(critical_fields)} field(s) have severe data gaps - consider removing or finding alternative data sources")
        if concerning_fields:
            recommendations.append(f"📋 {len(concerning_fields)} field(s) need attention - investigate why data is missing")
        if len(excellent_fields) == len(self.df.columns):
            recommendations.append("✅ All fields have excellent data completeness")
        
        return {
            'excellent_fields': excellent_fields,
            'good_fields': good_fields,
            'concerning_fields': concerning_fields,
            'critical_fields': critical_fields,
            'recommendations': recommendations,
            'overall_completeness': round((len(excellent_fields) + len(good_fields)) / len(self.df.columns) * 100, 1)
        }
    
    def _analyze_data_consistency(self) -> Dict[str, Any]:
        """Analyze data consistency patterns."""
        consistency_issues = []
        
        # Check for duplicate records
        duplicate_count = self.df.duplicated().sum()
        if duplicate_count > 0:
            consistency_issues.append(f"Found {duplicate_count:,} duplicate records that may cause double-counting")
        
        # Check for inconsistent formatting in text fields
        text_columns = self.df.select_dtypes(include=['object']).columns
        formatting_issues = []
        
        for col in text_columns:
            if self.df[col].dtype == 'object':
                # Check for mixed case issues
                non_null_values = self.df[col].dropna()
                if len(non_null_values) > 0:
                    has_mixed_case = any(str(val).islower() for val in non_null_values) and \
                                   any(str(val).isupper() for val in non_null_values)
                    if has_mixed_case:
                        formatting_issues.append(f"{col}: Mixed uppercase/lowercase formatting")
        
        return {
            'duplicate_records': duplicate_count,
            'duplicate_percentage': round((duplicate_count / len(self.df)) * 100, 2),
            'formatting_issues': formatting_issues,
            'consistency_score': self._calculate_consistency_score(duplicate_count, formatting_issues)
        }
    
    def _calculate_consistency_score(self, duplicates: int, formatting_issues: List[str]) -> float:
        """Calculate consistency score out of 100."""
        score = 100
        
        # Deduct for duplicates
        duplicate_penalty = min((duplicates / len(self.df)) * 50, 30)
        score -= duplicate_penalty
        
        # Deduct for formatting issues
        formatting_penalty = min(len(formatting_issues) * 5, 20)
        score -= formatting_penalty
        
        return max(round(score, 1), 0)
    
    def _analyze_data_accuracy(self) -> Dict[str, Any]:
        """Analyze data accuracy and unusual patterns."""
        accuracy_issues = []
        suspicious_patterns = []
        
        # Check numeric fields for outliers
        numeric_cols = self.df.select_dtypes(include=[np.number]).columns
        outlier_fields = []
        
        for col in numeric_cols:
            if len(self.df[col].dropna()) > 0:
                Q1 = self.df[col].quantile(0.25)
                Q3 = self.df[col].quantile(0.75)
                IQR = Q3 - Q1
                lower_bound = Q1 - 1.5 * IQR
                upper_bound = Q3 + 1.5 * IQR
                
                outliers = self.df[(self.df[col] < lower_bound) | (self.df[col] > upper_bound)]
                outlier_count = len(outliers)
                
                if outlier_count > 0:
                    outlier_percentage = (outlier_count / len(self.df)) * 100
                    outlier_fields.append({
                        'field': col,
                        'count': outlier_count,
                        'percentage': round(outlier_percentage, 1),
                        'severity': 'High' if outlier_percentage > 10 else 'Medium' if outlier_percentage > 5 else 'Low'
                    })
        
        # Check for impossible values
        for col in numeric_cols:
            if 'age' in col.lower() and self.df[col].max() > 150:
                accuracy_issues.append(f"{col}: Contains unrealistic age values (>150)")
            elif 'percentage' in col.lower() or 'percent' in col.lower():
                if self.df[col].max() > 100 or self.df[col].min() < 0:
                    accuracy_issues.append(f"{col}: Contains invalid percentage values")
        
        return {
            'outlier_fields': outlier_fields,
            'accuracy_issues': accuracy_issues,
            'suspicious_patterns': suspicious_patterns,
            'accuracy_score': self._calculate_accuracy_score(outlier_fields, accuracy_issues)
        }
    
    def _calculate_accuracy_score(self, outlier_fields: List[Dict], accuracy_issues: List[str]) -> float:
        """Calculate accuracy score out of 100."""
        score = 100
        
        # Deduct for outliers
        for field in outlier_fields:
            if field['severity'] == 'High':
                score -= 15
            elif field['severity'] == 'Medium':
                score -= 10
            else:
                score -= 5
        
        # Deduct for accuracy issues
        score -= len(accuracy_issues) * 10
        
        return max(round(score, 1), 0)
    
    def _analyze_data_reliability(self) -> Dict[str, Any]:
        """Analyze data reliability indicators."""
        reliability_indicators = []
        
        # Check for fields with very low variability
        numeric_cols = self.df.select_dtypes(include=[np.number]).columns
        low_variance_fields = []
        
        for col in numeric_cols:
            if len(self.df[col].dropna()) > 1:
                variance = self.df[col].var()
                if variance == 0:
                    low_variance_fields.append(f"{col}: All values are identical")
                elif self.df[col].nunique() == 1:
                    low_variance_fields.append(f"{col}: Only one unique value")
        
        # Check for fields with suspiciously high correlation
        high_correlations = []
        if len(numeric_cols) > 1:
            corr_matrix = self.df[numeric_cols].corr()
            for i in range(len(corr_matrix.columns)):
                for j in range(i+1, len(corr_matrix.columns)):
                    corr_value = corr_matrix.iloc[i, j]
                    if abs(corr_value) > 0.95 and not pd.isna(corr_value):
                        col1, col2 = corr_matrix.columns[i], corr_matrix.columns[j]
                        high_correlations.append({
                            'field1': col1,
                            'field2': col2,
                            'correlation': round(corr_value, 3),
                            'warning': 'These fields may contain duplicate information'
                        })
        
        return {
            'low_variance_fields': low_variance_fields,
            'high_correlations': high_correlations,
            'reliability_score': self._calculate_reliability_score(low_variance_fields, high_correlations)
        }
    
    def _calculate_reliability_score(self, low_variance: List[str], high_corr: List[Dict]) -> float:
        """Calculate reliability score out of 100."""
        score = 100
        
        # Deduct for low variance fields
        score -= len(low_variance) * 10
        
        # Deduct for high correlations
        score -= len(high_corr) * 8
        
        return max(round(score, 1), 0)
    
    def _assess_business_impact(self) -> Dict[str, Any]:
        """Assess business impact of data quality issues."""
        impact_assessment = {
            'overall_score': 0,
            'key_concerns': [],
            'action_priorities': [],
            'business_risks': []
        }
        
        # Calculate overall score
        scores = [
            self.report['overview']['data_health_score'],
            self.report['consistency']['consistency_score'],
            self.report['accuracy']['accuracy_score'],
            self.report['reliability']['reliability_score']
        ]
        impact_assessment['overall_score'] = round(sum(scores) / len(scores), 1)
        
        # Identify key concerns
        if self.report['completeness']['critical_fields']:
            impact_assessment['key_concerns'].append("Critical data gaps may affect analysis reliability")
        
        if self.report['consistency']['duplicate_records'] > len(self.df) * 0.05:
            impact_assessment['key_concerns'].append("High duplicate rate may skew results")
        
        if len(self.report['accuracy']['outlier_fields']) > 0:
            impact_assessment['key_concerns'].append("Unusual data patterns detected")
        
        # Set action priorities
        if impact_assessment['overall_score'] < 70:
            impact_assessment['action_priorities'] = [
                "1. Address critical data gaps immediately",
                "2. Remove or investigate duplicate records",
                "3. Validate unusual data patterns",
                "4. Implement data quality monitoring"
            ]
        elif impact_assessment['overall_score'] < 85:
            impact_assessment['action_priorities'] = [
                "1. Clean up minor data issues",
                "2. Standardize data formats",
                "3. Monitor for new quality issues"
            ]
        else:
            impact_assessment['action_priorities'] = [
                "1. Maintain current data quality standards",
                "2. Regular quality monitoring recommended"
            ]
        
        return impact_assessment
    
    def get_business_summary(self) -> str:
        """Get a business-friendly summary report."""
        overview = self.report['overview']
        completeness = self.report['completeness']
        impact = self.report['business_impact']
        
        summary_lines = [
            "DATA QUALITY HEALTH REPORT",
            "=" * 50,
            f"Dataset: {overview['total_records']:,} records across {overview['total_fields']} fields",
            f"Overall Health Score: {impact['overall_score']}/100 ({overview['health_status']})",
            f"Complete Records: {overview['complete_records']:,} ({round((overview['complete_records']/overview['total_records'])*100, 1)}%)",
            "",
            "FIELD QUALITY BREAKDOWN:",
            f"Excellent Quality: {len(completeness['excellent_fields'])} fields",
            f"Good Quality: {len(completeness['good_fields'])} fields",
            f"Needs Attention: {len(completeness['concerning_fields'])} fields",
            f"Critical Issues: {len(completeness['critical_fields'])} fields",
            "",
        ]
        
        # Add key concerns
        if impact['key_concerns']:
            summary_lines.append("KEY CONCERNS:")
            for concern in impact['key_concerns']:
                summary_lines.append(f"• {concern}")
            summary_lines.append("")
        
        # Add action priorities
        if impact['action_priorities']:
            summary_lines.append("RECOMMENDED ACTIONS:")
            for action in impact['action_priorities']:
                summary_lines.append(f"{action}")
            summary_lines.append("")
        
        # Add field details for critical issues
        if completeness['critical_fields']:
            summary_lines.append("FIELDS WITH CRITICAL ISSUES:")
            for field in completeness['critical_fields'][:5]:  # Show top 5
                missing_pct = round((self.df[field].isnull().sum() / len(self.df)) * 100, 1)
                summary_lines.append(f"• {field}: {missing_pct}% missing data")
            if len(completeness['critical_fields']) > 5:
                summary_lines.append(f"• ... and {len(completeness['critical_fields']) - 5} more fields")
            summary_lines.append("")
        
        return "\n".join(summary_lines)
    
    def export_detailed_report(self, file_path: str) -> bool:
        """Export detailed business report to Excel with highlighted issues."""
        try:
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create summary sheet
            summary_ws = wb.create_sheet("Executive Summary")
            self._create_summary_sheet(summary_ws)
            
            # Create data with highlights sheet
            data_ws = wb.create_sheet("Data with Issues Highlighted")
            self._create_highlighted_data_sheet(data_ws)
            
            # Create detailed analysis sheet
            analysis_ws = wb.create_sheet("Detailed Analysis")
            self._create_analysis_sheet(analysis_ws)
            
            # Save workbook
            wb.save(file_path)
            return True
            
        except Exception as e:
            print(f"Error exporting report: {e}")
            return False
    
    def _create_summary_sheet(self, ws):
        """Create executive summary sheet."""
        # Title
        ws['A1'] = "DATA QUALITY EXECUTIVE SUMMARY"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Key metrics
        row = 3
        overview = self.report['overview']
        impact = self.report['business_impact']
        
        metrics = [
            ("Total Records", f"{overview['total_records']:,}"),
            ("Total Fields", f"{overview['total_fields']}"),
            ("Overall Health Score", f"{impact['overall_score']}/100"),
            ("Health Status", overview['health_status']),
            ("Complete Records", f"{overview['complete_records']:,}"),
            ("Records with Issues", f"{overview['incomplete_records']:,}"),
        ]
        
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row = 1
        
        # Action items
        row += 2
        ws[f'A{row}'] = "RECOMMENDED ACTIONS"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        for i, action in enumerate(impact['action_priorities'], 1):
            ws[f'A{row}'] = action
            row += 1
    
    def _create_highlighted_data_sheet(self, ws):
        """Create data sheet with issues highlighted."""
        # Define colors for different issue types
        missing_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Light red
        outlier_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow
        duplicate_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # Light blue
        
        # Add headers
        for col_idx, column in enumerate(self.df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.font = Font(bold=True)
        
        # Add data with highlighting
        for row_idx, (_, row_data) in enumerate(self.df.iterrows(), 2):
            for col_idx, (column, value) in enumerate(row_data.items(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Highlight missing values
                if pd.isna(value):
                    cell.fill = missing_fill
                    cell.value = "MISSING DATA"
                
                # Highlight outliers for numeric columns
                elif column in self.df.select_dtypes(include=[np.number]).columns:
                    if not pd.isna(value):
                        Q1 = self.df[column].quantile(0.25)
                        Q3 = self.df[column].quantile(0.75)
                        IQR = Q3 - Q1
                        lower_bound = Q1 - 1.5 * IQR
                        upper_bound = Q3 + 1.5 * IQR
                        
                        if value < lower_bound or value > upper_bound:
                            cell.fill = outlier_fill
        
        # Highlight duplicate rows
        duplicate_rows = self.df[self.df.duplicated(keep=False)].index
        for row_idx in duplicate_rows:
            for col_idx in range(1, len(self.df.columns) + 1):
                cell = ws.cell(row=row_idx + 2, column=col_idx)  # +2 because Excel is 1-indexed and we have header
                if cell.fill.start_color.rgb != missing_fill.start_color.rgb and \
                   cell.fill.start_color.rgb != outlier_fill.start_color.rgb:
                    cell.fill = duplicate_fill
        
        # Add legend
        legend_row = len(self.df) + 3
        ws[f'A{legend_row}'] = "LEGEND:"
        ws[f'A{legend_row}'].font = Font(bold=True)
        
        legend_items = [
            ("Missing Data", missing_fill),
            ("Unusual Values", outlier_fill),
            ("Duplicate Records", duplicate_fill)
        ]
        
        for i, (label, fill) in enumerate(legend_items):
            cell = ws.cell(row=legend_row + 1 + i, column=1, value=label)
            cell.fill = fill
    
    def _create_analysis_sheet(self, ws):
        """Create detailed analysis sheet."""
        row = 1
        
        # Completeness analysis
        ws[f'A{row}'] = "DATA COMPLETENESS ANALYSIS"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        completeness = self.report['completeness']
        
        categories = [
            ("Excellent Fields (>95% complete)", completeness['excellent_fields']),
            ("Good Fields (85-95% complete)", completeness['good_fields']),
            ("Concerning Fields (50-85% complete)", completeness['concerning_fields']),
            ("Critical Fields (<50% complete)", completeness['critical_fields'])
        ]
        
        for category, fields in categories:
            ws[f'A{row}'] = f"{category}: {len(fields)}"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for field in fields[:10]:  # Show first 10
                missing_pct = round((self.df[field].isnull().sum() / len(self.df)) * 100, 1)
                ws[f'B{row}'] = f"{field}: {missing_pct}% missing"
                row += 1
            
            if len(fields) > 10:
                ws[f'B{row}'] = f"... and {len(fields) - 10} more fields"
                row += 1
            
            row += 1
        
        # Accuracy analysis
        row += 2
        ws[f'A{row}'] = "DATA ACCURACY ANALYSIS"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        accuracy = self.report['accuracy']
        if accuracy['outlier_fields']:
            ws[f'A{row}'] = "Fields with Unusual Values:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for field_info in accuracy['outlier_fields']:
                ws[f'B{row}'] = f"{field_info['field']}: {field_info['count']} unusual values ({field_info['percentage']}%) - {field_info['severity']} priority"
                row += 1


class BusinessDataQualityVisualizer:
    """Create business-friendly visualizations for data quality analysis."""
    
    def __init__(self, df: pd.DataFrame, analyzer: BusinessDataQualityAnalyzer):
        self.df = df
        self.analyzer = analyzer
        # Set business-friendly style
        plt.style.use('default')
        sns.set_palette("Set2")  # More business-friendly colors
    
    def create_health_dashboard(self) -> Figure:
        """Create overall health dashboard."""
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(15, 10))
        fig.suptitle('Data Quality Health Dashboard', fontsize=16, fontweight='bold')
        
        # Health score gauge
        self._create_health_gauge(ax1)
        
        # Field quality breakdown
        self._create_field_quality_chart(ax2)
        
        # Data completeness by field
        self._create_completeness_chart(ax3)
        
        # Issue severity breakdown
        self._create_issue_severity_chart(ax4)
        
        plt.tight_layout()
        return fig
    
    def _create_health_gauge(self, ax):
        """Create health score gauge chart."""
        score = self.analyzer.report['business_impact']['overall_score']
        
        # Create gauge
        theta = np.linspace(0, np.pi, 100)
        radius = 1
        
        # Background arc
        ax.plot(radius * np.cos(theta), radius * np.sin(theta), 'lightgray', linewidth=20)
        
        # Score arc
        score_theta = np.linspace(0, np.pi * (score / 100), int(score))
        if score >= 85:
            color = 'green'
        elif score >= 70:
            color = 'orange'
        else:
            color = 'red'
        
        ax.plot(radius * np.cos(score_theta), radius * np.sin(score_theta), color, linewidth=20)
        
        # Add score text
        ax.text(0, -0.3, f'{score}/100', ha='center', va='center', fontsize=20, fontweight='bold')
        ax.text(0, -0.5, 'Overall Health Score', ha='center', va='center', fontsize=12)
        
        ax.set_xlim(-1.2, 1.2)
        ax.set_ylim(-0.7, 1.2)
        ax.set_aspect('equal')
        ax.axis('off')
        ax.set_title('Data Health Score', fontweight='bold')
    
    def _create_field_quality_chart(self, ax):
        """Create field quality breakdown chart."""
        completeness = self.analyzer.report['completeness']
        
        categories = ['Excellent', 'Good', 'Needs Attention', 'Critical Issues']
        counts = [
            len(completeness['excellent_fields']),
            len(completeness['good_fields']),
            len(completeness['concerning_fields']),
            len(completeness['critical_fields'])
        ]
        colors = ['#2E8B57', '#32CD32', '#FFD700', '#FF6347']
        
        bars = ax.bar(categories, counts, color=colors)
        ax.set_title('Field Quality Breakdown', fontweight='bold')
        ax.set_ylabel('Number of Fields')
        
        # Add value labels on bars
        for bar in bars:
            height = bar.get_height()
            if height > 0:
                ax.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                       f'{int(height)}', ha='center', va='bottom', fontweight='bold')
        
        # Rotate x-axis labels for better readability
        plt.setp(ax.get_xticklabels(), rotation=45, ha='right')
    
    def _create_completeness_chart(self, ax):
        """Create data completeness chart."""
        missing_counts = self.df.isnull().sum()
        missing_percentages = (missing_counts / len(self.df)) * 100
        
        # Show top 10 fields with missing data
        top_missing = missing_percentages[missing_percentages > 0].sort_values(ascending=True).tail(10)
        
        if not top_missing.empty:
            colors = ['red' if x > 50 else 'orange' if x > 15 else 'yellow' for x in top_missing.values]
            bars = ax.barh(range(len(top_missing)), top_missing.values, color=colors)
            ax.set_yticks(range(len(top_missing)))
            ax.set_yticklabels(top_missing.index)
            ax.set_xlabel('Missing Data (%)')
            ax.set_title('Fields with Missing Data', fontweight='bold')
            
            # Add percentage labels
            for i, bar in enumerate(bars):
                width = bar.get_width()
                ax.text(width + 1, bar.get_y() + bar.get_height()/2,
                       f'{width:.1f}%', ha='left', va='center')
        else:
            ax.text(0.5, 0.5, 'No Missing Data Found!\n✅ All fields complete', 
                   transform=ax.transAxes, ha='center', va='center', 
                   fontsize=14, fontweight='bold', color='green')
            ax.set_title('Data Completeness Status', fontweight='bold')
            ax.axis('off')
    
    def _create_issue_severity_chart(self, ax):
        """Create issue severity breakdown."""
        # Count different types of issues
        completeness = self.analyzer.report['completeness']
        consistency = self.analyzer.report['consistency']
        accuracy = self.analyzer.report['accuracy']
        
        issue_counts = {
            'Critical Gaps': len(completeness['critical_fields']),
            'Duplicate Records': 1 if consistency['duplicate_records'] > 0 else 0,
            'Data Inconsistencies': len(consistency['formatting_issues']),
            'Unusual Values': len(accuracy['outlier_fields'])
        }
        
        # Filter out zero counts
        issue_counts = {k: v for k, v in issue_counts.items() if v > 0}
        
        if issue_counts:
            wedges, texts, autotexts = ax.pie(issue_counts.values(), labels=issue_counts.keys(), 
                                            autopct='%1.0f%%', startangle=90,
                                            colors=['#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A'])
            ax.set_title('Issue Types Distribution', fontweight='bold')
            
            # Make percentage text bold
            for autotext in autotexts:
                autotext.set_color('white')
                autotext.set_fontweight('bold')
        else:
            ax.text(0.5, 0.5, 'No Major Issues Detected!\n✅ Data looks good', 
                   transform=ax.transAxes, ha='center', va='center', 
                   fontsize=14, fontweight='bold', color='green')
            ax.set_title('Issue Analysis', fontweight='bold')
    
    def create_business_trends_chart(self) -> Figure:
        """Create business-relevant trend analysis."""
        fig, axes = plt.subplots(2, 2, figsize=(15, 10))
        fig.suptitle('Business Data Insights', fontsize=16, fontweight='bold')
        
        numeric_cols = self.df.select_dtypes(include=[np.number]).columns
        
        if len(numeric_cols) >= 2:
            # Top correlations
            ax1 = axes[0, 0]
            self._create_correlation_insights(ax1, numeric_cols)
            
            # Data distribution insights
            ax2 = axes[0, 1]
            self._create_distribution_insights(ax2, numeric_cols)
            
            # Value concentration analysis
            ax3 = axes[1, 0]
            self._create_concentration_analysis(ax3)
            
            # Data freshness indicators
            ax4 = axes[1, 1]
            self._create_freshness_indicators(ax4)
        else:
            fig.text(0.5, 0.5, 'Insufficient numeric data for trend analysis', 
                    ha='center', va='center', fontsize=16)
        
        plt.tight_layout()
        return fig
    
    def _create_correlation_insights(self, ax, numeric_cols):
        """Create business-friendly correlation insights."""
        if len(numeric_cols) < 2:
            ax.text(0.5, 0.5, 'Need at least 2 numeric fields\nfor correlation analysis', 
                   transform=ax.transAxes, ha='center', va='center')
            ax.set_title('Relationship Analysis')
            return
        
        # Calculate correlations
        corr_matrix = self.df[numeric_cols].corr()
        
        # Find strongest correlations (excluding self-correlations)
        correlations = []
        for i in range(len(corr_matrix.columns)):
            for j in range(i+1, len(corr_matrix.columns)):
                col1, col2 = corr_matrix.columns[i], corr_matrix.columns[j]
                corr_value = corr_matrix.iloc[i, j]
                if not pd.isna(corr_value):
                    correlations.append((abs(corr_value), col1, col2, corr_value))
        
        # Sort by strength and take top 5
        correlations.sort(reverse=True)
        top_correlations = correlations[:5]
        
        if top_correlations:
            labels = [f"{col1[:10]}...\n↔\n{col2[:10]}..." if len(col1) > 10 or len(col2) > 10 
                     else f"{col1}\n↔\n{col2}" for _, col1, col2, _ in top_correlations]
            values = [abs(corr) for abs_corr, _, _, corr in top_correlations]
            colors = ['green' if corr > 0 else 'red' for _, _, _, corr in top_correlations]
            
            bars = ax.barh(range(len(labels)), values, color=colors)
            ax.set_yticks(range(len(labels)))
            ax.set_yticklabels(labels)
            ax.set_xlabel('Relationship Strength')
            ax.set_title('Strongest Data Relationships', fontweight='bold')
            ax.set_xlim(0, 1)
            
            # Add correlation values
            for i, bar in enumerate(bars):
                width = bar.get_width()
                ax.text(width + 0.02, bar.get_y() + bar.get_height()/2,
                       f'{width:.2f}', ha='left', va='center', fontweight='bold')
        else:
            ax.text(0.5, 0.5, 'No significant relationships found', 
                   transform=ax.transAxes, ha='center', va='center')
        
        ax.set_title('Field Relationships', fontweight='bold')
    
    def _create_distribution_insights(self, ax, numeric_cols):
        """Create distribution insights."""
        if len(numeric_cols) == 0:
            ax.text(0.5, 0.5, 'No numeric fields found', 
                   transform=ax.transAxes, ha='center', va='center')
            ax.set_title('Value Distribution Analysis')
            return
        
        # Analyze skewness of numeric fields
        skewness_data = []
        for col in numeric_cols[:6]:  # Limit to 6 for readability
            skew = self.df[col].skew()
            if not pd.isna(skew):
                skewness_data.append((col, abs(skew)))
        
        if skewness_data:
            skewness_data.sort(key=lambda x: x[1], reverse=True)
            fields, skew_values = zip(*skewness_data)
            
            colors = ['red' if skew > 2 else 'orange' if skew > 1 else 'green' for skew in skew_values]
            bars = ax.bar(range(len(fields)), skew_values, color=colors)
            
            ax.set_xticks(range(len(fields)))
            ax.set_xticklabels([f[:8] + '...' if len(f) > 8 else f for f in fields], rotation=45)
            ax.set_ylabel('Distribution Skewness')
            ax.set_title('Data Distribution Analysis', fontweight='bold')
            
            # Add interpretation
            ax.axhline(y=1, color='orange', linestyle='--', alpha=0.7, label='Moderate Skew')
            ax.axhline(y=2, color='red', linestyle='--', alpha=0.7, label='High Skew')
            ax.legend()
    
    def _create_concentration_analysis(self, ax):
        """Analyze value concentration in categorical fields."""
        categorical_cols = self.df.select_dtypes(include=['object']).columns
        
        if len(categorical_cols) == 0:
            ax.text(0.5, 0.5, 'No categorical fields found', 
                   transform=ax.transAxes, ha='center', va='center')
            ax.set_title('Value Concentration Analysis')
            return
        
        # Calculate concentration (how many values make up 80% of data)
        concentration_data = []
        for col in categorical_cols[:5]:  # Limit to 5 for readability
            value_counts = self.df[col].value_counts()
            total_count = len(self.df)
            cumulative_pct = 0
            values_needed = 0
            
            for count in value_counts:
                cumulative_pct += (count / total_count)
                values_needed += 1
                if cumulative_pct >= 0.8:
                    break
            
            concentration_ratio = values_needed / len(value_counts) * 100
            concentration_data.append((col, concentration_ratio))
        
        if concentration_data:
            fields, concentrations = zip(*concentration_data)
            colors = ['green' if c < 20 else 'orange' if c < 50 else 'red' for c in concentrations]
            
            bars = ax.bar(range(len(fields)), concentrations, color=colors)
            ax.set_xticks(range(len(fields)))
            ax.set_xticklabels([f[:8] + '...' if len(f) > 8 else f for f in fields], rotation=45)
            ax.set_ylabel('Concentration (%)')
            ax.set_title('Value Concentration Analysis', fontweight='bold')
            
            # Add value labels
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height + 1,
                       f'{height:.0f}%', ha='center', va='bottom')
    
    def _create_freshness_indicators(self, ax):
        """Create data freshness indicators."""
        # Look for date columns
        date_cols = self.df.select_dtypes(include=['datetime64']).columns
        
        if len(date_cols) == 0:
            # Try to find date-like columns
            potential_date_cols = [col for col in self.df.columns 
                                 if any(keyword in col.lower() for keyword in ['date', 'time', 'created', 'updated'])]
            
            if potential_date_cols:
                ax.text(0.5, 0.5, f'Potential date fields found:\n{", ".join(potential_date_cols[:3])}\n\nConsider converting to date format\nfor freshness analysis', 
                       transform=ax.transAxes, ha='center', va='center', fontsize=10)
            else:
                ax.text(0.5, 0.5, 'No date fields found\n\nData freshness analysis\nnot available', 
                       transform=ax.transAxes, ha='center', va='center')
        else:
            # Analyze the most recent date column
            date_col = date_cols[0]
            dates = pd.to_datetime(self.df[date_col], errors='coerce').dropna()
            
            if len(dates) > 0:
                # Create date range histogram
                dates.hist(bins=20, ax=ax, alpha=0.7, color='skyblue')
                ax.set_title(f'Data Timeline - {date_col}', fontweight='bold')
                ax.set_xlabel('Date')
                ax.set_ylabel('Number of Records')
                
                # Add latest date annotation
                latest_date = dates.max()
                ax.axvline(latest_date, color='red', linestyle='--', label=f'Latest: {latest_date.strftime("%Y-%m-%d")}')
                ax.legend()
        
        ax.set_title('Data Freshness Analysis', fontweight='bold')


class InteractiveBusinessDataQuality:
    """Interactive business-friendly data quality reporting system."""
    
    def __init__(self, df: pd.DataFrame):
        self.df = df
        self.analyzer = BusinessDataQualityAnalyzer(df)
        self.visualizer = BusinessDataQualityVisualizer(df, self.analyzer)
        self.available_charts = {
            'health_dashboard': 'Data Health Dashboard',
            'business_trends': 'Business Insights & Trends'
        }
    
    def get_business_summary(self) -> str:
        """Get business-friendly summary report."""
        return self.analyzer.get_business_summary()
    
    def get_chart_options(self) -> Dict[str, str]:
        """Get available chart options."""
        return self.available_charts
    
    def create_chart(self, chart_type: str) -> Optional[Figure]:
        """Create a specific chart type."""
        try:
            if chart_type == 'health_dashboard':
                return self.visualizer.create_health_dashboard()
            elif chart_type == 'business_trends':
                return self.visualizer.create_business_trends_chart()
            else:
                return None
        except Exception as e:
            print(f"Error creating chart {chart_type}: {e}")
            return None
    
    def export_business_report(self, file_path: str) -> bool:
        """Export comprehensive business report with highlighted data issues."""
        return self.analyzer.export_detailed_report(file_path)
    
    def get_detailed_report(self) -> Dict[str, Any]:
        """Get detailed analysis report."""
        return self.analyzer.report
